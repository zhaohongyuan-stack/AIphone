"""
知识库 Excel 解析 + 拒绝解答检测
"""
import logging
from typing import Optional

from openpyxl import load_workbook

from config import settings

logger = logging.getLogger(__name__)


class KnowledgeBase:
    """
    知识库管理：加载 Excel，维护可回答/拒绝解答两组数据
    """

    def __init__(self):
        # 可回答列表: [{"question", "similar", "answer", "group"}, ...]
        self.answerable: list[dict] = []
        # 拒绝解答列表: [{"question", "similar", "answer", "group"}, ...]
        self.rejected: list[dict] = []
        # 拒绝解答关键词索引（用于快速匹配）
        self.reject_keywords: set[str] = set()
        self._load()

    def _load(self):
        """从 Excel 加载知识库"""
        try:
            wb = load_workbook(settings.KB_EXCEL_PATH, data_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                return
            # 表头: 分组名 | 问题 | 相似问法 | 答案
            for row in rows[1:]:
                if not row or not row[0]:
                    continue
                group, question, similar, answer = (
                    str(row[0] or "").strip(),
                    str(row[1] or "").strip(),
                    str(row[2] or "").strip(),
                    str(row[3] or "").strip(),
                )
                item = {"question": question, "similar": similar,
                        "answer": answer, "group": group}
                if "拒绝" in group:
                    self.rejected.append(item)
                    # 将问题和相似问法都拆成关键词
                    for kw in [question] + similar.split("\n"):
                        kw = kw.strip()
                        if kw:
                            self.reject_keywords.add(kw)
                else:
                    self.answerable.append(item)
            logger.info(
                f"知识库加载完成: 可回答 {len(self.answerable)} 条, "
                f"拒绝解答 {len(self.rejected)} 条"
            )
        except Exception as e:
            logger.error(f"知识库加载失败: {e}")

    def is_rejected(self, utterance: str) -> Optional[dict]:
        """
        检测用户提问是否命中"拒绝解答"
        命中则返回对应的拒绝项，否则返回 None
        """
        utt = utterance.strip()
        for item in self.rejected:
            # 精确匹配
            if utt == item["question"]:
                return item
            # 相似问法逐条匹配
            for kw in item["similar"].split("\n"):
                kw = kw.strip()
                if kw and kw in utt:
                    return item
        return None


# 全局单例
kb = KnowledgeBase()
